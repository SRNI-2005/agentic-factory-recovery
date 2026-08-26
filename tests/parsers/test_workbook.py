import io

import pytest
from openpyxl import load_workbook

from coe.parsers.workbook import (
    SHEETS,
    WorkbookRejected,
    apply_workbook,
    export_workbook,
    validate_workbook,
)

pytestmark = pytest.mark.db


def test_export_contains_all_sheets_with_headers(clean_db, session,
                                                 demo_scenario):
    blob = export_workbook(session, demo_scenario)
    wb = load_workbook(io.BytesIO(blob))
    assert set(SHEETS) <= set(wb.sheetnames)
    for name, headers in SHEETS.items():
        ws = wb[name]
        assert [c.value for c in ws[1]] == list(headers)


def test_export_jobs_match_database(clean_db, session, demo_scenario):
    from coe.db.models.fjsp import Job

    blob = export_workbook(session, demo_scenario)
    ws = load_workbook(io.BytesIO(blob))["Jobs"]
    exported = {r[0] for r in ws.iter_rows(min_row=2, values_only=True)}
    db_names = {n for (n,) in session.query(Job.name)
                .filter(Job.instance_id == demo_scenario).all()}
    assert exported == db_names


def test_meta_sheet_records_source(clean_db, session, demo_scenario):
    blob = export_workbook(session, demo_scenario)
    ws = load_workbook(io.BytesIO(blob))["Meta"]
    meta = {r[0]: r[1] for r in ws.iter_rows(min_row=2, values_only=True)}
    assert meta["exported_from"] == "factory_demo_01"
    assert "target_name" in meta


def test_export_bom_matches_database(clean_db, session, demo_scenario):
    from coe.db.models.fjsp import Job, Operation
    from coe.db.models.materials import Material, OperationBom

    blob = export_workbook(session, demo_scenario)
    ws = load_workbook(io.BytesIO(blob))["BOM"]
    exported = set(ws.iter_rows(min_row=2, values_only=True))
    op_meta = {o.id: (j.name, o.sequence_number)
               for o, j in session.query(Operation, Job)
               .join(Job, Operation.job_id == Job.id)
               .filter(Operation.instance_id == demo_scenario)}
    sku_of = {m.id: m.sku for m in session.query(Material)
              .filter(Material.instance_id == demo_scenario)}
    db_rows = {(op_meta[b.operation_id][0], op_meta[b.operation_id][1],
                sku_of[b.material_id], b.quantity_required)
               for b in session.query(OperationBom)
               .filter(OperationBom.instance_id == demo_scenario)}
    assert db_rows
    assert exported == db_rows


def _edit(blob: bytes, sheet: str, mutate) -> bytes:
    """Load export bytes, apply mutate(wb), return new bytes."""
    wb = load_workbook(io.BytesIO(blob))
    mutate(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parent(session):
    from coe.db.models.provenance import Instance

    return (session.query(Instance)
            .filter(Instance.name == "factory_demo_01").one())


def ws_row_of_append(instance_id: int) -> int:
    """Row number the appended bad receipt occupied (2 + existing rows)."""
    from coe.db.models.materials import MaterialReceipt

    from coe.db.session import session_scope
    with session_scope() as s:
        n = s.query(MaterialReceipt).filter(
            MaterialReceipt.instance_id == instance_id).count()
    return n + 2


def _add_scheduled_job(session, instance_id: int, job_name: str):
    """Job+op+alternative+one schedule entry, to prove removal guard."""
    from coe.db.models.fjsp import (
        Job, Machine, Operation, OperationMachineAlternative)
    from coe.db.models.schedule import ScheduleEntry, ScheduleVersion

    mach = (session.query(Machine)
            .filter(Machine.instance_id == instance_id)
            .order_by(Machine.name).first())
    job = Job(instance_id=instance_id, name=job_name, release_time=0,
              deadline=500, priority=1, status="PENDING")
    session.add(job)
    session.flush()
    op = Operation(instance_id=instance_id, job_id=job.id,
                   sequence_number=1)
    session.add(op)
    session.flush()
    session.add(OperationMachineAlternative(
        instance_id=instance_id, operation_id=op.id, machine_id=mach.id,
        processing_time=10))
    ver = ScheduleVersion(instance_id=instance_id, version_number=999,
                          schedule_type="BASELINE", solver_status="FEASIBLE",
                          objective_value=1.0, makespan=10, total_tardiness=0,
                          alpha_weight=0.5, beta_weight=0.5,
                          time_limit_seconds=1, solve_duration_seconds=0.1,
                          failed_machine_ids=None, parent_version_id=None,
                          rolled_back=False, payload_hash="0" * 64,
                          payload_json={})
    session.add(ver)
    session.flush()
    session.add(ScheduleEntry(instance_id=instance_id, version_id=ver.id,
                              operation_id=op.id, machine_id=mach.id,
                              worker_id=None, start_time=0, end_time=10,
                              processing_time=10, setup_time=0,
                              status="SCHEDULED", is_frozen=False))
    session.flush()


def test_missing_column_rejected(clean_db, session, demo_scenario):
    blob = export_workbook(session, demo_scenario)

    def drop_col(wb):
        ws = wb["Jobs"]
        ws.delete_cols(3)          # release_time

    errs = validate_workbook(_edit(blob, "Jobs", drop_col), session,
                             _parent(session))
    assert any("missing column" in e["message"] and e["sheet"] == "Jobs"
               for e in errs)


def test_unknown_sku_rejected_with_row_number(clean_db, session,
                                              demo_scenario):
    blob = export_workbook(session, demo_scenario)

    def bad_receipt(wb):
        ws = wb["Receipts"]
        ws.append(("MAT-NOPE", 5, 100, "test"))

    errs = validate_workbook(_edit(blob, "Receipts", bad_receipt), session,
                             _parent(session))
    hit = [e for e in errs if e["sheet"] == "Receipts" and "MAT-NOPE" in
           e["message"]]
    assert hit and hit[0]["row"] == ws_row_of_append(demo_scenario)


def test_negative_processing_time_rejected(clean_db, session, demo_scenario):
    blob = export_workbook(session, demo_scenario)

    def negate(wb):
        ws = wb["Alternatives"]
        ws.cell(row=2, column=4, value=-5)

    errs = validate_workbook(_edit(blob, "Alternatives", negate), session,
                             _parent(session))
    assert any("non-negative" in e["message"] for e in errs)


def test_speed_without_alternative_rejected(clean_db, session,
                                            demo_scenario):
    from coe.db.models.fjsp import Job, Machine
    from coe.db.models.workers import Worker

    blob = export_workbook(session, demo_scenario)
    jname = session.query(Job.name).filter(
        Job.instance_id == demo_scenario).order_by(Job.name).first()[0]
    mname = session.query(Machine.name).filter(
        Machine.instance_id == demo_scenario).order_by(Machine.name).first()[0]
    wname = session.query(Worker.name).filter(
        Worker.instance_id == demo_scenario).order_by(Worker.name).first()[0]

    def add_speed(wb):
        wb["Speeds"].append((jname, 99, mname, wname, 7))

    errs = validate_workbook(_edit(blob, "Speeds", add_speed), session,
                             _parent(session))
    assert any("not in Alternatives" in e["message"] for e in errs)


def test_scheduled_job_removal_rejected(clean_db, session, demo_scenario):
    """Removing a scheduled job's rows must point the user at Suspend."""
    _add_scheduled_job(session, demo_scenario, "ZZZ-sched")
    blob = export_workbook(session, demo_scenario)

    def remove_zzz(wb):
        ws = wb["Jobs"]
        for r in list(ws.iter_rows(min_row=2)):
            if r[0].value == "ZZZ-sched":
                ws.delete_rows(r[0].row)

    errs = validate_workbook(_edit(blob, "Jobs", remove_zzz), session,
                             _parent(session))
    assert any("Suspend" in e["message"] for e in errs)


def test_unknown_setup_machine_rejected(clean_db, session, demo_scenario):
    blob = export_workbook(session, demo_scenario)

    def ghost_machine(wb):
        wb["Setups"].append(("MACHINE-GHOST", None, None, 5))

    errs = validate_workbook(_edit(blob, "Setups", ghost_machine), session,
                             _parent(session))
    assert any(e["sheet"] == "Setups" and "unknown machine" in e["message"]
               for e in errs)


def test_duplicate_speed_rejected(clean_db, session, demo_scenario):
    blob = export_workbook(session, demo_scenario)
    first = load_workbook(io.BytesIO(blob))["Speeds"]
    row2 = tuple(c.value for c in first[2])

    def dup(wb):
        wb["Speeds"].append(row2)

    errs = validate_workbook(_edit(blob, "Speeds", dup), session,
                             _parent(session))
    assert any("duplicate speed" in e["message"] for e in errs)


def test_pristine_export_validates_clean(clean_db, session, demo_scenario):
    blob = export_workbook(session, demo_scenario)
    assert validate_workbook(blob, session, _parent(session)) == []


def test_new_material_with_receipt_validates(clean_db, session,
                                             demo_scenario):
    """Coordinated same-file Materials+Receipts addition must not be rejected."""
    blob = export_workbook(session, demo_scenario)

    def add_pair(wb):
        wb["Materials"].append(("MAT-NEW", 10, 0))
        wb["Receipts"].append(("MAT-NEW", 5, 100, "test"))

    errs = validate_workbook(_edit(blob, "Materials", add_pair), session,
                             _parent(session))
    assert errs == []


def _semantic_state(session, iid: int) -> dict[str, set]:
    """Id-free multiset snapshot of every editable domain."""
    from sqlalchemy import text

    def q(sql):
        return {tuple(r) for r in session.execute(text(sql),
                                                  {"i": iid}).all()}

    return {
        "Jobs": q(
            "SELECT j.name, COALESCE(f.name,''), j.release_time, "
            "       j.deadline, j.priority FROM jobs j "
            "LEFT JOIN job_families f ON f.id = j.job_family_id "
            "WHERE j.instance_id = :i"),
        "Alternatives": q(
            "SELECT j.name, o.sequence_number, m.name, "
            "       a.processing_time FROM operation_machine_alternatives a "
            "JOIN operations o ON o.id = a.operation_id "
            "JOIN jobs j ON j.id = o.job_id "
            "JOIN machines m ON m.id = a.machine_id "
            "WHERE a.instance_id = :i"),
        "Speeds": q(
            "SELECT j.name, o.sequence_number, m.name, w.name, "
            "       t.processing_time FROM operation_machine_worker_times t "
            "JOIN operations o ON o.id = t.operation_id "
            "JOIN jobs j ON j.id = o.job_id "
            "JOIN machines m ON m.id = t.machine_id "
            "JOIN workers w ON w.id = t.worker_id "
            "WHERE t.instance_id = :i"),
        "Setups": q(
            "SELECT m.name, ff.name, tt.name, s.setup_duration "
            "FROM setup_times s JOIN machines m ON m.id = s.machine_id "
            "LEFT JOIN job_families ff ON ff.id = s.from_family_id "
            "LEFT JOIN job_families tt ON tt.id = s.to_family_id "
            "WHERE s.instance_id = :i"),
        "Materials": q(
            "SELECT sku, initial_stock, reorder_point FROM materials "
            "WHERE instance_id = :i"),
        "Receipts": q(
            "SELECT mat.sku, r.quantity, r.available_at, r.source "
            "FROM material_receipts r JOIN materials mat ON mat.id = "
            "r.material_id WHERE r.instance_id = :i"),
        "Availability": q(
            "SELECT w.name, v.available_from, v.available_until "
            "FROM worker_availability_windows v "
            "JOIN workers w ON w.id = v.worker_id "
            "WHERE v.instance_id = :i"),
        "BOM": q(
            "SELECT j.name, o.sequence_number, mat.sku, b.quantity_required "
            "FROM operation_bom b JOIN operations o ON o.id = b.operation_id "
            "JOIN jobs j ON j.id = o.job_id "
            "JOIN materials mat ON mat.id = b.material_id "
            "WHERE b.instance_id = :i"),
    }


def test_round_trip_identity(clean_db, session, demo_scenario):
    """export → apply unchanged ⇒ every editable domain semantically equal."""
    blob = export_workbook(session, demo_scenario)
    fork = apply_workbook(session, _parent(session), blob,
                          new_name="rt-check")
    before = _semantic_state(session, demo_scenario)
    after = _semantic_state(session, fork.id)
    assert before == after


def test_round_trip_keeps_schedule_fk_valid(clean_db, session,
                                            demo_scenario):
    _add_scheduled_job(session, demo_scenario, "ZZZ-sched")
    blob = export_workbook(session, demo_scenario)
    fork = apply_workbook(session, _parent(session), blob,
                          new_name="rt-sched")
    from sqlalchemy import text

    broken = session.execute(text(
        "SELECT count(*) FROM schedule_entries se "
        "WHERE se.instance_id = :f AND NOT EXISTS ("
        "  SELECT 1 FROM operations o WHERE o.id = se.operation_id)"
    ), {"f": fork.id}).scalar_one()
    assert broken == 0
    # and the ZZZ entry still points at an op named ZZZ-sched#1 equivalent
    mapped = session.execute(text(
        "SELECT j.name, o.sequence_number FROM schedule_entries se "
        "JOIN operations o ON o.id = se.operation_id "
        "JOIN jobs j ON j.id = o.job_id WHERE se.instance_id = :f"
    ), {"f": fork.id}).all()
    assert [(n, sq) for (n, sq) in mapped] == [("ZZZ-sched", 1)]


def test_edits_landing_and_parent_pristine(clean_db, session,
                                           demo_scenario):
    blob = export_workbook(session, demo_scenario)

    def bump_and_add(wb):
        ws = wb["Materials"]
        ws.cell(row=2, column=2, value=ws.cell(row=2, column=2).value + 7)
        wb["Jobs"].append(("NEWJOB", None, 0, 300, 1))
        wb["Alternatives"].append(
            ("NEWJOB", 1,
             wb["Alternatives"].cell(row=2, column=3).value, 12))

    fork = apply_workbook(session, _parent(session),
                          _edit(blob, "Materials", bump_and_add),
                          new_name="edited")
    # sku -> full row (state rows are (sku, stock, reorder) triples)
    mats = {r[0]: r for r in _semantic_state(session, fork.id)["Materials"]}
    parent_mats = _semantic_state(session, demo_scenario)["Materials"]
    first_sku = sorted(parent_mats)[0][0]
    assert mats[first_sku][1] == sorted(
        parent_mats)[0][1] + 7                      # stock bumped by 7
    assert ("NEWJOB", "", 0, 300, 1) in \
        _semantic_state(session, fork.id)["Jobs"]   # job added
    # parent untouched:
    assert _semantic_state(session, demo_scenario)["Materials"] \
        == parent_mats


def test_invalid_apply_raises_without_instance(clean_db, session,
                                               demo_scenario):
    from coe.db.models.provenance import Instance

    blob = export_workbook(session, demo_scenario)

    def poison(wb):
        wb["Receipts"].append(("MAT-GHOST", 5, 100, "test"))

    n_before = session.query(Instance).count()
    with pytest.raises(WorkbookRejected):
        apply_workbook(session, _parent(session),
                       _edit(blob, "Receipts", poison))
    assert session.query(Instance).count() == n_before
