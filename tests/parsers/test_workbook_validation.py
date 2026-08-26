"""Failing tests for final-fixes-brief Critical #1 and Important #5.

Required integer cells (release_time, op_sequence, processing_time, etc.)
must produce row-level validation errors when blank, NOT crash in apply_workbook
with int(None).  Required string identifiers (sku, machine, worker) must
likewise be caught as row-level errors.
"""
import io

import pytest
from openpyxl import Workbook, load_workbook

from coe.parsers.workbook import (
    SHEETS,
    WorkbookRejected,
    apply_workbook,
    export_workbook,
    validate_workbook,
)

pytestmark = pytest.mark.db


def _edit(blob: bytes, sheet: str, mutate) -> bytes:
    wb = load_workbook(io.BytesIO(blob))
    mutate(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _rebuild_with_blank(blob: bytes, sheet: str, row_idx: int,
                        col_idx: int) -> bytes:
    """Reload export, blank one cell by rebuilding row without it."""
    wb = load_workbook(io.BytesIO(blob))
    ws = wb[sheet]
    # Read all rows, blank the target cell
    rows = []
    for r in ws.iter_rows(min_row=1, values_only=True):
        rows.append(list(r))
    # rows[0] = header, rows[row_idx] = target (1-indexed in openpyxl)
    rows[row_idx][col_idx - 1] = None
    # Clear and rewrite
    for r in list(ws.iter_rows(min_row=1)):
        ws.delete_rows(1)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parent(session):
    from coe.db.models.provenance import Instance
    return (session.query(Instance)
            .filter(Instance.name == "factory_demo_01").one())


# ------------------------------------------------------------------ #
# Critical #1: required integer cells must error when blank            #
# ------------------------------------------------------------------ #

class TestBlankIntegerValidation:
    """Blank mandatory integer columns must produce row-level errors."""

    def test_blank_release_time_rejected(self, clean_db, session,
                                         demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "Jobs", 2, 3)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("release_time" in e["message"]
                   and e["sheet"] == "Jobs" for e in errs), \
            f"expected release_time blank error, got {errs}"

    def test_blank_op_sequence_rejected(self, clean_db, session,
                                        demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "Alternatives", 2, 2)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("op_sequence" in e["message"]
                   and e["sheet"] == "Alternatives" for e in errs), \
            f"expected op_sequence blank error, got {errs}"

    def test_blank_processing_time_rejected(self, clean_db, session,
                                            demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "Alternatives", 2, 4)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("processing_time" in e["message"]
                   and e["sheet"] == "Alternatives" for e in errs), \
            f"expected processing_time blank error, got {errs}"

    def test_blank_priority_rejected(self, clean_db, session, demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "Jobs", 2, 5)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("priority" in e["message"]
                   and e["sheet"] == "Jobs" for e in errs), \
            f"expected priority blank error, got {errs}"

    def test_blank_initial_stock_rejected(self, clean_db, session,
                                          demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "Materials", 2, 2)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("initial_stock" in e["message"]
                   and e["sheet"] == "Materials" for e in errs), \
            f"expected initial_stock blank error, got {errs}"

    def test_blank_quantity_rejected(self, clean_db, session, demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "Receipts", 2, 2)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("quantity" in e["message"]
                   and e["sheet"] == "Receipts" for e in errs), \
            f"expected quantity blank error, got {errs}"

    def test_blank_setup_duration_rejected(self, clean_db, session,
                                           demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "Setups", 2, 4)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("setup_duration" in e["message"]
                   and e["sheet"] == "Setups" for e in errs), \
            f"expected setup_duration blank error, got {errs}"

    def test_blank_available_from_rejected(self, clean_db, session,
                                           demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "Availability", 2, 2)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("available_from" in e["message"]
                   and e["sheet"] == "Availability" for e in errs), \
            f"expected available_from blank error, got {errs}"

    def test_blank_quantity_required_rejected(self, clean_db, session,
                                              demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "BOM", 2, 4)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("quantity_required" in e["message"]
                   and e["sheet"] == "BOM" for e in errs), \
            f"expected quantity_required blank error, got {errs}"


# ------------------------------------------------------------------ #
# Important #5: required string identifiers blank → row error          #
# ------------------------------------------------------------------ #

class TestBlankStringIdentifierValidation:
    """Blank required string identifiers must produce row-level errors."""

    def test_blank_material_sku_rejected(self, clean_db, session,
                                         demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "Materials", 2, 1)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("sku" in e["message"].lower()
                   and e["sheet"] == "Materials" for e in errs), \
            f"expected sku blank error, got {errs}"

    def test_blank_alternative_machine_rejected(self, clean_db, session,
                                                demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "Alternatives", 2, 3)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("machine" in e["message"].lower()
                   and e["sheet"] == "Alternatives" for e in errs), \
            f"expected machine blank error, got {errs}"

    def test_blank_speed_machine_rejected(self, clean_db, session,
                                          demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "Speeds", 2, 3)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("machine" in e["message"].lower()
                   and e["sheet"] == "Speeds" for e in errs), \
            f"expected machine blank error in Speeds, got {errs}"

    def test_blank_speed_worker_rejected(self, clean_db, session,
                                         demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "Speeds", 2, 4)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("worker" in e["message"].lower()
                   and e["sheet"] == "Speeds" for e in errs), \
            f"expected worker blank error in Speeds, got {errs}"

    def test_blank_setup_machine_rejected(self, clean_db, session,
                                          demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "Setups", 2, 1)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("machine" in e["message"].lower()
                   and e["sheet"] == "Setups" for e in errs), \
            f"expected machine blank error in Setups, got {errs}"

    def test_blank_availability_worker_rejected(self, clean_db, session,
                                                demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "Availability", 2, 1)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("worker" in e["message"].lower()
                   and e["sheet"] == "Availability" for e in errs), \
            f"expected worker blank error in Availability, got {errs}"

    def test_blank_bom_job_rejected(self, clean_db, session, demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "BOM", 2, 1)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("job" in e["message"].lower()
                   and e["sheet"] == "BOM" for e in errs), \
            f"expected job blank error in BOM, got {errs}"

    def test_blank_bom_sku_rejected(self, clean_db, session, demo_scenario):
        blob = export_workbook(session, demo_scenario)
        blob = _rebuild_with_blank(blob, "BOM", 2, 3)
        errs = validate_workbook(blob, session, _parent(session))
        assert any("sku" in e["message"].lower()
                   and e["sheet"] == "BOM" for e in errs), \
            f"expected sku blank error in BOM, got {errs}"
