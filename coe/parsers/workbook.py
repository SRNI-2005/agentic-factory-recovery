"""Fourth importer: user-authored factory workbook (dashboard design §3).

Editable domains only; physics tables are inherited verbatim from the
parent instance via fork_instance(). Import is two-phase: validate_workbook()
produces row-level errors without writing; apply_workbook() forks then
replaces covered domains atomically.
"""
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

SHEETS: dict[str, tuple[str, ...]] = {
    "Meta": ("key", "value"),
    "Jobs": ("name", "family", "release_time", "deadline", "priority"),
    "Alternatives": ("job", "op_sequence", "machine", "processing_time"),
    "Speeds": ("job", "op_sequence", "machine", "worker",
               "processing_time"),
    "Setups": ("machine", "from_family", "to_family", "setup_duration"),
    "Materials": ("sku", "initial_stock", "reorder_point"),
    "Receipts": ("sku", "quantity", "available_at", "source"),
    "Availability": ("worker", "available_from", "available_until"),
    "BOM": ("job", "op_sequence", "sku", "quantity_required"),
}

_INT_COLS = {"op_sequence", "release_time", "deadline", "priority",
             "processing_time", "setup_duration", "initial_stock",
             "reorder_point", "quantity", "available_at",
             "available_from", "available_until", "quantity_required"}

WorkbookError = dict  # {"sheet": str, "row": int | None, "message": str}


class WorkbookRejected(ValueError):
    def __init__(self, errors: list[dict]):
        self.errors = errors
        lines = "\n".join(f"[{e['sheet']}#{e['row']}] {e['message']}"
                          for e in errors[:20])
        super().__init__(f"workbook rejected, {len(errors)} problem(s):\n{lines}")


def _sort_key(row: tuple) -> tuple:
    # Setup families may be NULL (family-agnostic setups); None must not be
    # compared against str during row ordering.
    return tuple((v is not None, v) for v in row)


def _sheet_rows(session: Session, model, instance_id: int,
                order_col, project):
    rows = (session.query(model)
            .filter(model.instance_id == instance_id)
            .order_by(order_col).all())
    return [project(r) for r in rows]


def export_workbook(session: Session, instance_id: int) -> bytes:
    """Serialize the editable domains of an instance to xlsx bytes."""
    from coe.db.models.fjsp import (
        Job, JobFamily, Machine, Operation, OperationMachineAlternative,
        SetupTime)
    from coe.db.models.materials import Material, MaterialReceipt, OperationBom
    from coe.db.models.provenance import Instance
    from coe.db.models.workers import (
        OperationMachineWorkerTime, Worker, WorkerAvailabilityWindow)

    inst = session.get(Instance, instance_id)
    fam = {f.id: f.name for f in session.query(JobFamily)
           .filter(JobFamily.instance_id == instance_id)}
    mach = {m.id: m.name for m in session.query(Machine)
            .filter(Machine.instance_id == instance_id)}
    work = {w.id: w.name for w in session.query(Worker)
            .filter(Worker.instance_id == instance_id)}
    jobs = {j.id: j.name for j in session.query(Job)
            .filter(Job.instance_id == instance_id)}
    fam_by_name = {v: k for k, v in fam.items()}
    mach_by_name = {v: k for k, v in mach.items()}

    data: dict[str, list[tuple]] = {}
    data["Jobs"] = _sheet_rows(
        session, Job, instance_id, Job.name,
        lambda r: (r.name, fam.get(r.job_family_id), r.release_time,
                   r.deadline, r.priority))
    ops = (session.query(Operation)
           .filter(Operation.instance_id == instance_id)
           .order_by(Operation.id).all())
    op_key = {(jobs[o.job_id], o.sequence_number): o.id for o in ops}
    alt_rows = _sheet_rows(
        session, OperationMachineAlternative, instance_id,
        OperationMachineAlternative.operation_id,
        lambda r: (*[k for k, v in op_key.items() if v == r.operation_id][0],
                   mach[r.machine_id], r.processing_time))
    data["Alternatives"] = sorted(alt_rows)
    data["Speeds"] = sorted(_sheet_rows(
        session, OperationMachineWorkerTime, instance_id,
        OperationMachineWorkerTime.operation_id,
        lambda r: (*[k for k, v in op_key.items() if v == r.operation_id][0],
                   mach[r.machine_id], work[r.worker_id],
                   r.processing_time)))
    data["Setups"] = sorted(_sheet_rows(
        session, SetupTime, instance_id, SetupTime.id,
        lambda r: (mach[r.machine_id],
                   fam.get(r.from_family_id), fam.get(r.to_family_id),
                   r.setup_duration)), key=_sort_key)
    data["Materials"] = _sheet_rows(
        session, Material, instance_id, Material.sku,
        lambda r: (r.sku, r.initial_stock, r.reorder_point))
    mat = {m.sku: m.id for m in session.query(Material)
           .filter(Material.instance_id == instance_id)}
    data["Receipts"] = sorted(_sheet_rows(
        session, MaterialReceipt, instance_id, MaterialReceipt.id,
        lambda r: (*[k for k, v in mat.items() if v == r.material_id],
                   r.quantity, r.available_at, r.source)))
    data["Availability"] = sorted(_sheet_rows(
        session, WorkerAvailabilityWindow, instance_id,
        WorkerAvailabilityWindow.id,
        lambda r: (work[r.worker_id], r.available_from,
                   r.available_until)))
    data["BOM"] = sorted(_sheet_rows(
        session, OperationBom, instance_id, OperationBom.operation_id,
        lambda r: (*[k for k, v in op_key.items() if v == r.operation_id][0],
                   *[k for k, v in mat.items() if v == r.material_id],
                   r.quantity_required)))

    wb = Workbook()
    wb.remove(wb.active)
    meta = wb.create_sheet("Meta")
    meta.append(("key", "value"))
    meta.append(("exported_from", inst.name))
    meta.append(("exported_at",
                 datetime.now(timezone.utc).isoformat(timespec="seconds")))
    meta.append(("target_name", f"{inst.name}-edited"))
    for name, headers in SHEETS.items():
        if name == "Meta":
            continue
        ws = wb.create_sheet(name)
        ws.append(list(headers))
        for row in data.get(name, []):
            ws.append(list(row))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def load_rows(data: bytes) -> dict[str, list[dict]]:
    """Parse xlsx into per-sheet dict rows after header verification."""
    wb = load_workbook(BytesIO(data), data_only=True)
    problems: list[dict] = []
    for name in SHEETS:
        if name not in wb.sheetnames:
            problems.append({"sheet": name, "row": None,
                             "message": "missing sheet"})
    if problems:
        raise WorkbookRejected(problems)
    out: dict[str, list[dict]] = {}
    for name, headers in SHEETS.items():
        ws = wb[name]
        actual = [c.value for c in ws[1]]
        for h in headers:
            if h not in actual:
                problems.append({"sheet": name, "row": 1,
                                 "message": f"missing column '{h}'"})
        out[name] = []
        for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True),
                                   start=2):
            if all(v is None for v in row):
                continue
            out[name].append(dict(zip(actual, row)) | {"_row": ridx})
    if problems:
        raise WorkbookRejected(problems)
    return out


def _as_int(sheet: str, row: dict, col: str, errors: list[dict],
            *, positive: bool = False, nonneg: bool = True) -> int | None:
    raw = row.get(col)
    if raw is None:
        return None
    try:
        val = int(raw)
    except (TypeError, ValueError):
        errors.append({"sheet": sheet, "row": row["_row"],
                       "message": f"'{col}' must be an integer, got {raw!r}"})
        return None
    if positive and val <= 0:
        errors.append({"sheet": sheet, "row": row["_row"],
                       "message": f"'{col}' must be > 0, got {val}"})
    elif nonneg and val < 0:
        errors.append({"sheet": sheet, "row": row["_row"],
                       "message": f"'{col}' must be non-negative, got {val}"})
    return val


def validate_workbook(data: bytes, session: Session, parent) -> list[dict]:
    """Full dry-run. Returns [] when the workbook is applicable.

    Name references resolve against the PARENT instance: apply forks the
    parent first, so parent scope == fork scope for inherited tables.
    """
    from coe.db.models.fjsp import Job, JobFamily, Machine, Operation
    from coe.db.models.materials import Material
    from coe.db.models.schedule import ScheduleEntry
    from coe.db.models.workers import Worker

    try:
        rows = load_rows(data)
    except WorkbookRejected as exc:
        return exc.errors

    errors: list[dict] = []
    pid = parent.id

    fam_names = {n for (n,) in session.query(JobFamily.name)
                 .filter(JobFamily.instance_id == pid)}
    mach_names = {n for (n,) in session.query(Machine.name)
                  .filter(Machine.instance_id == pid)}
    work_names = {n for (n,) in session.query(Worker.name)
                  .filter(Worker.instance_id == pid)}
    sku_names = {n for (n,) in session.query(Material.sku)
                 .filter(Material.instance_id == pid)}
    job_names = {n for (n,) in session.query(Job.name)
                 .filter(Job.instance_id == pid)}

    jobs_in_file: set[str] = set()
    seen_job_rows: set[str] = set()
    for r in rows["Jobs"]:
        name = r.get("name")
        if not name:
            errors.append({"sheet": "Jobs", "row": r["_row"],
                           "message": "job name required"})
            continue
        if name in seen_job_rows:
            errors.append({"sheet": "Jobs", "row": r["_row"],
                           "message": f"duplicate job '{name}'"})
        seen_job_rows.add(name)
        jobs_in_file.add(name)
        fam = r.get("family")
        if fam is not None and fam not in fam_names:
            errors.append({"sheet": "Jobs", "row": r["_row"],
                           "message": f"unknown family '{fam}'"})
        _as_int("Jobs", r, "release_time", errors)
        _as_int("Jobs", r, "deadline", errors)
        pr = _as_int("Jobs", r, "priority", errors)
        if pr is not None and pr < 1:
            errors.append({"sheet": "Jobs", "row": r["_row"],
                           "message": f"priority must be >= 1, got {pr}"})

    alt_keys: set[tuple] = set()
    seen_alt: set[tuple] = set()
    for r in rows["Alternatives"]:
        j, sq, m = r.get("job"), _as_int("Alternatives", r, "op_sequence",
                                         errors), r.get("machine")
        pt = _as_int("Alternatives", r, "processing_time", errors)
        key = (j, sq, m)
        if key in seen_alt:
            errors.append({"sheet": "Alternatives", "row": r["_row"],
                           "message": f"duplicate alternative {key}"})
        seen_alt.add(key)
        if j is not None:
            alt_keys.add((j, sq))
        if m is not None and m not in mach_names:
            errors.append({"sheet": "Alternatives", "row": r["_row"],
                           "message": f"unknown machine '{m}'"})
        if pt is not None and pt < 0:
            pass  # already reported by _as_int non-negative check

    seen_speed: set[tuple] = set()
    for r in rows["Speeds"]:
        j = r.get("job")
        sq = _as_int("Speeds", r, "op_sequence", errors)
        m, w = r.get("machine"), r.get("worker")
        _as_int("Speeds", r, "processing_time", errors)
        key = (j, sq, m, w)
        if key in seen_speed:
            errors.append({"sheet": "Speeds", "row": r["_row"],
                           "message": f"duplicate speed {key}"})
        seen_speed.add(key)
        if m not in mach_names:
            errors.append({"sheet": "Speeds", "row": r["_row"],
                           "message": f"unknown machine '{m}'"})
        if w not in work_names:
            errors.append({"sheet": "Speeds", "row": r["_row"],
                           "message": f"unknown worker '{w}'"})
        if (j, sq) not in alt_keys:
            errors.append({"sheet": "Speeds", "row": r["_row"],
                           "message":
                           f"speed for ('{j}',{sq}) not in Alternatives"})

    seen_setup: set[tuple] = set()
    for r in rows["Setups"]:
        m = r.get("machine")
        ff, tt = r.get("from_family"), r.get("to_family")
        if m is not None and m not in mach_names:
            errors.append({"sheet": "Setups", "row": r["_row"],
                           "message": f"unknown machine '{m}'"})
        _as_int("Setups", r, "setup_duration", errors, positive=True)
        for fname, label in ((ff, "from_family"), (tt, "to_family")):
            if fname is not None and fname not in fam_names:
                errors.append({"sheet": "Setups", "row": r["_row"],
                               "message": f"unknown {label} '{fname}'"})
        key = (m, ff, tt)
        if key in seen_setup:
            errors.append({"sheet": "Setups", "row": r["_row"],
                           "message": f"duplicate setup {key}"})
        seen_setup.add(key)

    seen_sku: set[str] = set()
    for r in rows["Materials"]:
        sku = r.get("sku")
        _as_int("Materials", r, "initial_stock", errors, nonneg=True)
        _as_int("Materials", r, "reorder_point", errors)
        if sku in seen_sku:
            errors.append({"sheet": "Materials", "row": r["_row"],
                           "message": f"duplicate sku '{sku}'"})
        seen_sku.add(sku or "")

    for r in rows["Receipts"]:
        if (r.get("sku") not in sku_names
                and r.get("sku") not in seen_sku):
            errors.append({"sheet": "Receipts", "row": r["_row"],
                           "message":
                           f"unknown material '{r.get('sku')}' — "
                           "define it under Materials"})
        q = _as_int("Receipts", r, "quantity", errors, positive=True)
        _as_int("Receipts", r, "available_at", errors)

    for r in rows["Availability"]:
        if r.get("worker") not in work_names:
            errors.append({"sheet": "Availability", "row": r["_row"],
                           "message": f"unknown worker '{r.get('worker')}'"})
        a = _as_int("Availability", r, "available_from", errors)
        b = _as_int("Availability", r, "available_until", errors)
        if a is not None and b is not None and b <= a:
            errors.append({"sheet": "Availability", "row": r["_row"],
                           "message":
                           f"available_until ({b}) must exceed "
                           f"available_from ({a})"})

    for r in rows["BOM"]:
        j = r.get("job")
        sq = _as_int("BOM", r, "op_sequence", errors)
        _as_int("BOM", r, "quantity_required", errors, positive=True)
        if r.get("sku") not in sku_names and r.get("sku") not in seen_sku:
            errors.append({"sheet": "BOM", "row": r["_row"],
                           "message": f"unknown material '{r.get('sku')}'"})
        if (j, sq) not in alt_keys:
            errors.append({"sheet": "BOM", "row": r["_row"],
                           "message":
                           f"BOM for ('{j}',{sq}) not in Alternatives"})

    # ---- removal guards against DB state ----
    id_to_job = {j.id: j.name for j in session.query(Job)
                 .filter(Job.instance_id == pid)}
    existing_ops = {(id_to_job[o.job_id], o.sequence_number)
                    for o in session.query(Operation)
                    .filter(Operation.instance_id == pid)}

    removed_jobs = job_names - jobs_in_file
    shrunk_ops = existing_ops - alt_keys
    if removed_jobs or shrunk_ops:
        guarded_ids = {e.operation_id for e in
                       session.query(ScheduleEntry.operation_id)
                       .filter(ScheduleEntry.instance_id == pid)}
        guarded = {(id_to_job[o.job_id], o.sequence_number)
                   for o in session.query(Operation)
                   .filter(Operation.instance_id == pid,
                           Operation.id.in_(guarded_ids))}
        for job_name in sorted(removed_jobs):
            if any(jn == job_name for jn, _ in guarded):
                errors.append({
                    "sheet": "Jobs", "row": None,
                    "message":
                    f"job '{job_name}' has committed schedule entries — "
                    "remove it via the Suspend action instead"})
        for key in sorted(shrunk_ops):
            if key in guarded:
                errors.append({
                    "sheet": "Alternatives", "row": None,
                    "message":
                    f"operation {key} has committed schedule entries — "
                    "restore it in Alternatives or Suspend its job"})
    return errors


def planned_job_names(rows: dict) -> set[str]:
    return {r["name"] for r in rows["Jobs"]}


def apply_workbook(session: Session, parent, data: bytes,
                   new_name: str | None = None):
    """Validate → fork → replace covered domains (single transaction).

    Deletion order matters: dependents (bom/speeds/alts/setups/receipts/
    availability) die before materials; vanished OPERATIONS die before
    their vanished JOB (the validation guard proved both unscheduled).
    """
    from coe.services.fork import fork_instance
    from coe.db.models.fjsp import (
        Job,
        JobFamily,
        Machine,
        Operation,
        OperationMachineAlternative,
        SetupTime,
    )
    from coe.db.models.materials import (
        Material,
        MaterialReceipt,
        OperationBom,
    )
    from coe.db.models.workers import (
        OperationMachineWorkerTime,
        Worker,
        WorkerAvailabilityWindow,
    )

    errors = validate_workbook(data, session, parent)
    if errors:
        raise WorkbookRejected(errors)
    rows = load_rows(data)

    if new_name is None:
        meta = {r.get("key"): r.get("value") for r in rows["Meta"]}
        new_name = meta.get("target_name")

    fork = fork_instance(session, parent, new_name=new_name)
    fid = fork.id

    fam = {f.name: f.id for f in session.query(JobFamily)
           .filter(JobFamily.instance_id == fid)}
    mach = {m.name: m.id for m in session.query(Machine)
            .filter(Machine.instance_id == fid)}
    work = {w.name: w.id for w in session.query(Worker)
            .filter(Worker.instance_id == fid)}

    planned_alt_keys = {(r["job"], int(r["op_sequence"]))
                        for r in rows["Alternatives"]}
    wanted_jobs = planned_job_names(rows)

    # --- snapshot current ops keyed by (job_name, seq) ---
    job_name_by_id = {j.id: j.name for j in session.query(Job)
                      .filter(Job.instance_id == fid)}
    old_jobs = {j.name: j for j in session.query(Job)
                .filter(Job.instance_id == fid)}
    old_ops = {}
    for o in session.query(Operation).filter(Operation.instance_id == fid):
        old_ops[(job_name_by_id[o.job_id], o.sequence_number)] = o

    # --- wipe replaced domains, FK-safe order ---
    def wipe(model):
        session.query(model).filter(model.instance_id == fid) \
            .delete(synchronize_session=False)

    wipe(OperationBom)
    wipe(OperationMachineWorkerTime)
    wipe(OperationMachineAlternative)
    wipe(SetupTime)
    wipe(MaterialReceipt)
    wipe(WorkerAvailabilityWindow)
    wipe(Material)
    session.flush()

    # --- prune vanished operations FIRST (unscheduled per guard),
    #     then their vanished jobs ---
    for key, op in list(old_ops.items()):
        if key not in planned_alt_keys:
            session.delete(op)
            del old_ops[key]
    session.flush()
    for name, job in list(old_jobs.items()):
        if name not in wanted_jobs:
            session.delete(job)
            del old_jobs[name]
    session.flush()

    # --- upsert jobs ---
    for r in rows["Jobs"]:
        vals = {"release_time": int(r["release_time"]),
                "deadline": (int(r["deadline"])
                             if r.get("deadline") is not None else None),
                "priority": int(r["priority"]),
                "job_family_id": (fam[r["family"]]
                                  if r.get("family") is not None else None)}
        if r["name"] in old_jobs:
            for k, v in vals.items():
                setattr(old_jobs[r["name"]], k, v)
        else:
            nj = Job(instance_id=fid, name=r["name"], status="PENDING",
                     **vals)
            session.add(nj)
            old_jobs[r["name"]] = nj
    session.flush()

    # --- insert missing operations (kept ops retain ids ⇒ entries valid) ---
    for key in sorted(planned_alt_keys - set(old_ops)):
        jname, seq = key
        old_ops[key] = Operation(instance_id=fid,
                                 job_id=old_jobs[jname].id,
                                 sequence_number=seq, required_role_id=None)
        session.add(old_ops[key])
    session.flush()

    # --- reinsert replaced domains from sheets ---
    for r in rows["Materials"]:
        session.add(Material(
            instance_id=fid, sku=r["sku"],
            initial_stock=int(r["initial_stock"]),
            reorder_point=(int(r["reorder_point"])
                           if r.get("reorder_point") is not None else None)))
    session.flush()
    sku_ids = {m.sku: m.id for m in session.query(Material)
               .filter(Material.instance_id == fid)}

    for r in rows["Receipts"]:
        session.add(MaterialReceipt(
            instance_id=fid, material_id=sku_ids[r["sku"]],
            quantity=int(r["quantity"]),
            available_at=int(r["available_at"]),
            source=r.get("source") or "workbook"))
    for r in rows["Availability"]:
        session.add(WorkerAvailabilityWindow(
            instance_id=fid, worker_id=work[r["worker"]],
            available_from=int(r["available_from"]),
            available_until=int(r["available_until"]),
            source_pattern="workbook"))
    for r in rows["Setups"]:
        session.add(SetupTime(
            instance_id=fid, machine_id=mach[r["machine"]],
            from_family_id=(fam[r["from_family"]]
                            if r.get("from_family") is not None else None),
            to_family_id=(fam[r["to_family"]]
                          if r.get("to_family") is not None else None),
            setup_duration=int(r["setup_duration"]), source="workbook"))
    for r in rows["Alternatives"]:
        op = old_ops[(r["job"], int(r["op_sequence"]))]
        session.add(OperationMachineAlternative(
            instance_id=fid, operation_id=op.id,
            machine_id=mach[r["machine"]],
            processing_time=int(r["processing_time"])))
    for r in rows["Speeds"]:
        op = old_ops[(r["job"], int(r["op_sequence"]))]
        session.add(OperationMachineWorkerTime(
            instance_id=fid, operation_id=op.id,
            machine_id=mach[r["machine"]], worker_id=work[r["worker"]],
            processing_time=int(r["processing_time"])))
    for r in rows["BOM"]:
        op = old_ops[(r["job"], int(r["op_sequence"]))]
        session.add(OperationBom(instance_id=fid, operation_id=op.id,
                                 material_id=sku_ids[r["sku"]],
                                 quantity_required=int(r[
                                     "quantity_required"])))
    session.flush()
    return fork
