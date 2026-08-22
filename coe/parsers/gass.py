from pathlib import Path

from openpyxl import load_workbook

from coe.db.models.provenance import InstanceProfile
from coe.db.session import session_scope
from coe.parsers.common import (
    SourceParseError,
    get_or_create_source_instance,
    sha256_file,
)

EXPECTED_FILES = (
    "1-Machine.xlsx", "2-Process.xlsx", "3-Routing.xlsx",
    "4-Width.xlsx", "5-Product Type.xlsx", "6-Data Order.xlsx",
)


def parse_manifest(path: Path) -> dict[str, str]:
    """manifest.txt lines: '<filename>|<uuid>|<sha256>'."""
    out: dict[str, str] = {}
    for lineno, line in enumerate(path.read_text().splitlines(), start=1):
        if not line.strip():
            continue
        parts = line.split("|")
        if len(parts) != 3:
            raise SourceParseError(f"{path.name}:{lineno}: expected 3 '|'-separated fields")
        fname, _, sha = parts
        out[fname.strip()] = sha.strip()
    return out


def _rows(sheet):
    """Yield value-tuples after the two header rows, skipping blank rows."""
    iterator = sheet.iter_rows(values_only=True)
    next(iterator, None)
    next(iterator, None)
    for row in iterator:
        if any(v is not None for v in row):
            yield row


def _extract(data_dir: Path) -> list[InstanceProfile]:
    def sheet(name: str):
        wb = load_workbook(data_dir / name, read_only=True, data_only=True)
        ws = wb.active
        rows = list(_rows(ws))
        wb.close()
        return rows

    machine_rows = sheet("1-Machine.xlsx")
    machines = [
        {
            "code": r[2],
            "name": r[1],
            "min_speed": int(r[4]),
            "ratio_speed": float(r[5]),
            "setup_time": int(r[6]),
        }
        for r in machine_rows
        if r[1] is not None
    ]

    process_rows = sheet("2-Process.xlsx")
    processes = [{"code": r[0], "name": r[1]} for r in process_rows if r[0] is not None]

    routing_rows = sheet("3-Routing.xlsx")
    routings = [
        {"id": int(r[0]), "sequence": str(r[1]).strip().split("-")}
        for r in routing_rows
        if r[0] is not None and r[1]
    ]

    width_wb = load_workbook(data_dir / "4-Width.xlsx", read_only=True, data_only=True)
    width_ws = width_wb["4-Width"]
    film_widths = [int(r[2]) for r in _rows(width_ws) if r[2] is not None]
    width_wb.close()

    type_rows = sheet("5-Product Type.xlsx")
    product_types = [
        {"width_mm": int(r[1]), "colors": int(r[2]), "routing_id": int(r[3])}
        for r in type_rows
        if r[1] is not None
    ]

    order_rows = sheet("6-Data Order.xlsx")
    orders = [
        {
            "no": int(r[0]),
            "priority": int(r[1]),
            "product_type": int(r[2]),
            "running_meter": int(r[3]),
            "lead_days": int(r[6]) if r[6] is not None else 14,
        }
        for r in order_rows
        if r[0] is not None
    ]

    return [
        InstanceProfile(
            name="gass-machines",
            profile_type="machine_setup",
            parameters_json={"machines": machines},
        ),
        InstanceProfile(
            name="gass-routings",
            profile_type="routing",
            parameters_json={
                "processes": processes,
                "routings": routings,
                "film_widths": film_widths,
                "product_types": product_types,
            },
        ),
        InstanceProfile(
            name="gass-orders",
            profile_type="order_pattern",
            parameters_json={"orders": orders},
        ),
    ]


def import_gass(data_dir: Path, instance_name: str = "gass") -> int:
    manifest = parse_manifest(data_dir / "manifest.txt")
    mismatches = [
        f"{fname}: expected {sha}, got {sha256_file(data_dir / fname)}"
        for fname, sha in manifest.items()
        if sha256_file(data_dir / fname) != sha
    ]
    if mismatches:
        raise SourceParseError("GASS checksum mismatch(es):\n" + "\n".join(mismatches))
    for fname in EXPECTED_FILES:
        if not (data_dir / fname).exists():
            raise SourceParseError(f"missing GASS file: {fname}")

    profiles = _extract(data_dir)
    with session_scope() as session:
        inst, created = get_or_create_source_instance(
            session,
            name=instance_name,
            source_name="gass-flexible-packaging",
            source_url=None,  # no verified public URL; see Task 8 note
            source_version="released-xlsx",
            source_license="academic-benchmark",
            checksum=sha256_file(data_dir / "manifest.txt"),
        )
        if not created:
            return inst.id
        for p in profiles:
            p.source_instance_id = inst.id
            session.add(p)
        session.flush()
        return inst.id
